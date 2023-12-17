import logging
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import time
import tkinter as tk
from tkinter import ttk
import os
import tkinter
import tkinter.filedialog
import tkinter.messagebox
import threading
import openpyxl
import configparser


def clicked():
    thread1 = threading.Thread(target=function)
    thread1.start()


def get_text2(driver, class_name):
    element = driver.find_element_by_class_name(class_name)
    text = element.get_attribute("textContent")
    return text


def num_make(num_str):
    if '万' in num_str:
        num_str = num_str[:-1]
        num_str = float(num_str)*10000
    elif '億' in num_str:
        num_str = num_str[:-1]
        num_str = float(num_str) * 100000000
    else:
        try:
            num_str = float(num_str)
        except:
            print('型エラー')
    return num_str


def function():
    root = tkinter.Tk()
    root.withdraw()
    fTyp = [("", "*")]
    iDir = os.path.abspath(os.path.dirname(__file__))
    # tkinter.messagebox.showinfo('○×プログラム','処理ファイルを選択してください！')
    file = tkinter.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
    # file = '入力.xlsx'
    wb1 = openpyxl.load_workbook(file)
    ws1 = wb1.worksheets[0]
    ws1.cell(1, 1).value = 'id'
    ws1.cell(1, 2).value = 'アカウント名'
    ws1.cell(1, 3).value = 'フォロー数'
    ws1.cell(1, 4).value = 'フォロワー数'
    ws1.cell(1, 5).value = '投稿数'
    url_account_list = []
    k = 1
    for cell in ws1['A']:
        if k > 1:
            url_account_pre = 'https://www.instagram.com/' + cell.value + '/?hl=ja'
            url_account_list.append(url_account_pre)
        k = k + 1
    url_all_num = len(url_account_list)
    #initial_url = 'https://www.instagram.com/kantei/?hl=ja'
    options = Options()
    '''
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--proxy-server="direct://"')
    options.add_argument('--proxy-bypass-list=*')
    options.add_argument('--start-maximized')
    '''
    options.add_argument('--no-sandbox')
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--lang=ja-JP')
    options.add_argument(
        f'user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.79 Safari/537.36')  # 追加

    #driver = webdriver.Chrome(executable_path="C:\\Users\\tisk0\\PycharmProjects\\chromedriver_win32\\chromedriver.exe",options=options)
    driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
    # driver.maximize_window()
    #driver.set_window_size('1200', '1000')
    driver.implicitly_wait(10)

    save_name.set('読み込み中')

    # try:
    # ログイン
    r_config = configparser.ConfigParser()
    r_config.read('Config.ini')
    ID = r_config.get('login', 'id')
    PASSWORD = r_config.get('login', 'password')
    driver.get('https://www.instagram.com/')
    time.sleep(2)
    # driver.save_screenshot('エラー確認.png')
    id_element = driver.find_element_by_css_selector('#loginForm > div > div:nth-child(1) > div > label > input')
    id_element.send_keys(ID)
    password_element = driver.find_element_by_css_selector('#loginForm > div > div:nth-child(2) > div > label > input')
    password_element.send_keys(PASSWORD)
    button_element = driver.find_element_by_css_selector('#loginForm > div > div:nth-child(3)')
    button_element.click()
    time.sleep(5)
    ct_url = 1
    for url_account in url_account_list:
        try:
            driver.get(url_account)
            #name = get_text2(driver,'rhpdm')
            #name = get_text2(driver,'_7UhW9       fKFbl yUEEX   KV-D4              fDxYl     ')
            #name_element = driver.find_element_by_css_selector('._7UhW9.fKFbl.yUEEX.KV-D4.fDxYl.')
            #name_element = driver.find_element_by_xpath('//div[@class="_7UhW9       fKFbl yUEEX   KV-D4              fDxYl     "]')
            name_element_pre = driver.find_element_by_class_name('XBGH5')
            name_element = name_element_pre.find_element_by_tag_name('h2')
            name = name_element.get_attribute("textContent")
            num_elements = driver.find_elements_by_class_name('g47SY ')
            post_num = num_elements[0].get_attribute("textContent")
            follow_num = num_elements[1].get_attribute("textContent")
            follower_num = num_elements[2].get_attribute("textContent")
            ws1.cell(ct_url + 1, 2).value = name
            ws1.cell(ct_url + 1, 3).value = num_make(follower_num)
            ws1.cell(ct_url + 1, 4).value = num_make(follow_num)
            ws1.cell(ct_url + 1, 5).value = num_make(post_num)
            # print(name)
            # print(post_num)
            # print(follower_num)
            # print(follow_num)
        except:
            # print('取得不可')
            logging.exception("What is doing when exception happens.")
            ws1.cell(ct_url + 1, 2).value = '取得不可'
            ws1.cell(ct_url + 1, 3).value = '取得不可'
            ws1.cell(ct_url + 1, 4).value = '取得不可'
            ws1.cell(ct_url + 1, 5).value = '取得不可'

        # time.sleep(1)
        save_name.set(str(ct_url) + '/' + str(url_all_num) + '件完了')
        ct_url = ct_url + 1

    save_name.set('全' + str(url_all_num) + '件完了')
    wb1.save(file)

    # except:
    # save_name.set('エラー発生')


root = tk.Tk()
root.title('instagramツール')
frame2 = ttk.Frame(root, padding=16)
button1 = ttk.Button(
    frame2,
    text='ファイルを選択',
    command=clicked)
frame4 = ttk.Frame(root, padding=16)
save_name = tk.StringVar()
entry4 = ttk.Entry(frame4, textvariable=save_name, width=30)
frame2.pack(side=tk.TOP, anchor=tk.NW)
button1.pack(fill=tk.X)
frame4.pack(side=tk.TOP, anchor=tk.NW)
entry4.pack(side=tk.TOP, anchor=tk.NW)

# ウィンドウの表示開始
root.mainloop()
